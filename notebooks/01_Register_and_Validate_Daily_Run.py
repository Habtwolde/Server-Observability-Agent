# Databricks notebook source
# 01_Register_and_Validate_Daily_Run

from __future__ import annotations

import hashlib
import json
import os
import re
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from pyspark.sql.types import (
    ArrayType,
    BooleanType,
    DateType,
    IntegerType,
    LongType,
    StringType,
    StructField,
    StructType,
    TimestampType,
)


# -----------------------------------------------------------------------------
# 1. Runtime parameters
# -----------------------------------------------------------------------------

dbutils.widgets.text("run_date", "", "Run date (YYYY-MM-DD; blank = today)")
dbutils.widgets.dropdown("bootstrap_registry", "false", ["false", "true"], "Bootstrap 45-server registry")
dbutils.widgets.dropdown("fail_on_incomplete", "false", ["false", "true"], "Fail when run is incomplete")
dbutils.widgets.text("run_trigger", "MANUAL", "Run trigger")

CATALOG = "ent_log_analytics"
SCHEMA = "observability"
VOLUME = "server_observability_vol"


def table_name(name: str) -> str:
    return f"`{CATALOG}`.`{SCHEMA}`.`{name}`"


config_rows = spark.table(f"{CATALOG}.{SCHEMA}.agent_config").select(
    "config_key", "config_value"
).collect()
CONFIG = {row["config_key"]: row["config_value"] for row in config_rows}

EXPECTED_SERVER_COUNT = int(CONFIG.get("expected_server_count", "45"))
SOURCE_TIMEZONE = CONFIG.get("source_timezone", "America/New_York")
SQL_INBOX = CONFIG["sql_diagnostics_inbox_path"]
SQL_BY_SERVER = CONFIG["sql_diagnostics_by_server_path"]
WINDOWS_INBOX = CONFIG["windows_events_inbox_path"]
AGENT_ROOT = CONFIG["agent_root_path"]

RUN_DATE_PARAMETER = dbutils.widgets.get("run_date").strip()
BOOTSTRAP_REGISTRY = dbutils.widgets.get("bootstrap_registry").strip().lower() == "true"
FAIL_ON_INCOMPLETE = dbutils.widgets.get("fail_on_incomplete").strip().lower() == "true"
RUN_TRIGGER = dbutils.widgets.get("run_trigger").strip().upper() or "MANUAL"

source_tz = ZoneInfo(SOURCE_TIMEZONE)
now_utc = datetime.now(timezone.utc)
now_local = now_utc.astimezone(source_tz)

if RUN_DATE_PARAMETER:
    target_run_date = date.fromisoformat(RUN_DATE_PARAMETER)
else:
    target_run_date = now_local.date()

print(f"Run date: {target_run_date}")
print(f"Source timezone: {SOURCE_TIMEZONE}")
print(f"Expected SQL workbooks: {EXPECTED_SERVER_COUNT}")
print(f"Bootstrap registry: {BOOTSTRAP_REGISTRY}")


# -----------------------------------------------------------------------------
# 2. File and workbook helpers
# -----------------------------------------------------------------------------

SUPPORTED_WORKBOOK_EXTENSIONS = {".xlsx", ".xlsm"}
SUPPORTED_WINDOWS_EXTENSIONS = {".csv"}


def to_fuse_path(path: str) -> str:
    """Convert dbfs:/Volumes/... into the local /Volumes/... FUSE path."""
    if path.startswith("dbfs:/Volumes/"):
        return path[len("dbfs:"):]
    return path


def normalize_server_name(value: Any) -> str:
    value_text = "" if value is None else str(value)
    normalized = re.sub(r"\s+", "", value_text).strip().upper()
    return normalized


def safe_path_component(value: str) -> str:
    safe = re.sub(r"[^A-Z0-9._$-]+", "_", value.upper()).strip("._")
    return safe or "UNKNOWN_SERVER"


def filename_server_candidate(file_name: str) -> str:
    stem = Path(file_name).stem
    stem = re.sub(r"\(\d+\)$", "", stem).strip()
    return normalize_server_name(stem)


def sha256_file(path: str, chunk_size: int = 4 * 1024 * 1024) -> str:
    digest = hashlib.sha256()
    with open(to_fuse_path(path), "rb") as handle:
        while True:
            block = handle.read(chunk_size)
            if not block:
                break
            digest.update(block)
    return digest.hexdigest()


def source_file_id(
    source_type: str,
    source_path: str,
    modification_time_ms: int,
    size_bytes: int,
    content_sha256: str,
) -> str:
    observation = "|".join(
        [
            source_type,
            source_path,
            str(modification_time_ms),
            str(size_bytes),
            content_sha256,
        ]
    )
    return hashlib.sha256(observation.encode("utf-8")).hexdigest()


def modification_timestamp(modification_time_ms: int) -> datetime:
    return datetime.fromtimestamp(modification_time_ms / 1000, tz=timezone.utc)


def list_current_files(folder: str, extensions: set[str]) -> tuple[list[Any], list[Any]]:
    current_files = []
    stale_files = []

    for item in dbutils.fs.ls(folder):
        if item.name.endswith("/") or item.name.startswith("~$"):
            continue

        if Path(item.name).suffix.lower() not in extensions:
            continue

        modified_utc = modification_timestamp(item.modificationTime)
        modified_local = modified_utc.astimezone(source_tz)

        if modified_local.date() == target_run_date:
            current_files.append(item)
        else:
            stale_files.append(item)

    return current_files, stale_files


SERVER_HEADER_ALIASES = {
    "servername",
    "sqlservername",
    "serverinstance",
    "sqlserverinstance",
}


def extract_workbook_identity(path: str) -> dict[str, Any]:
    """
    Read the first worksheet only.

    The verified production layout is:
      Sheet: 1-Version Info
      A1: Server Name
      A2: <server identifier>

    A limited label search is retained for minor collector layout changes.
    """
    workbook = None

    try:
        workbook = load_workbook(
            filename=to_fuse_path(path),
            read_only=True,
            data_only=True,
        )

        sheet_names = list(workbook.sheetnames)
        if not sheet_names:
            raise ValueError("Workbook has no worksheets")

        first_sheet_name = sheet_names[0]
        worksheet = workbook[first_sheet_name]

        exact_header = worksheet["A1"].value
        exact_value = worksheet["A2"].value

        if normalize_server_name(exact_header) == "SERVERNAME" and exact_value:
            reported_server = normalize_server_name(exact_value)
        else:
            reported_server = ""

            for row_number in range(1, 21):
                for column_number in range(1, 11):
                    label_value = worksheet.cell(
                        row=row_number,
                        column=column_number,
                    ).value

                    normalized_label = re.sub(
                        r"[^a-z0-9]",
                        "",
                        "" if label_value is None else str(label_value).lower(),
                    )

                    if normalized_label not in SERVER_HEADER_ALIASES:
                        continue

                    below_value = worksheet.cell(
                        row=row_number + 1,
                        column=column_number,
                    ).value

                    right_value = worksheet.cell(
                        row=row_number,
                        column=column_number + 1,
                    ).value

                    candidate = below_value or right_value
                    if candidate:
                        reported_server = normalize_server_name(candidate)
                        break

                if reported_server:
                    break

        if not reported_server:
            raise ValueError(
                "Server identity was not found in the first worksheet"
            )

        if not re.fullmatch(r"[A-Z0-9][A-Z0-9._$\\-]{0,127}", reported_server):
            raise ValueError(
                f"Workbook server identity has unsupported characters: {reported_server}"
            )

        return {
            "workbook_reported_server": reported_server,
            "first_sheet_name": first_sheet_name,
            "sheet_count": len(sheet_names),
            "validation_error": None,
        }

    except Exception as exc:
        return {
            "workbook_reported_server": None,
            "first_sheet_name": None,
            "sheet_count": None,
            "validation_error": str(exc),
        }

    finally:
        if workbook is not None:
            workbook.close()


def path_exists(path: str) -> bool:
    return os.path.exists(to_fuse_path(path))


# -----------------------------------------------------------------------------
# 3. Discover today's SQL workbooks and Windows Events file
# -----------------------------------------------------------------------------

sql_files, stale_sql_files = list_current_files(
    SQL_INBOX,
    SUPPORTED_WORKBOOK_EXTENSIONS,
)

windows_files, stale_windows_files = list_current_files(
    WINDOWS_INBOX,
    SUPPORTED_WINDOWS_EXTENSIONS,
)

print(f"Current SQL workbooks discovered: {len(sql_files)}")
print(f"Stale SQL workbooks ignored: {len(stale_sql_files)}")
print(f"Current Windows Events files discovered: {len(windows_files)}")
print(f"Stale Windows Events files ignored: {len(stale_windows_files)}")

file_records: list[dict[str, Any]] = []

for file_info in sql_files:
    content_hash = sha256_file(file_info.path)
    current_source_file_id = source_file_id(
        "SQL_DIAGNOSTICS",
        file_info.path,
        file_info.modificationTime,
        file_info.size,
        content_hash,
    )

    identity = extract_workbook_identity(file_info.path)
    reported_server = identity["workbook_reported_server"]
    candidate_from_filename = filename_server_candidate(file_info.name)
    modified_utc = modification_timestamp(file_info.modificationTime)

    if identity["validation_error"]:
        status = "INVALID_WORKBOOK"
        validation_message = identity["validation_error"]
    else:
        status = "VALIDATED"

        if reported_server not in candidate_from_filename:
            validation_message = (
                "Workbook identity is authoritative. Filename does not contain "
                f"the reported server name: workbook={reported_server}, "
                f"filename_candidate={candidate_from_filename}."
            )
        else:
            validation_message = (
                f"Validated from {identity['first_sheet_name']}; "
                f"worksheet_count={identity['sheet_count']}."
            )

    file_records.append(
        {
            "source_file_id": current_source_file_id,
            "run_id": None,
            "source_type": "SQL_DIAGNOSTICS",
            "original_file_name": file_info.name,
            "inbox_file_path": file_info.path,
            "archived_file_path": None,
            "content_sha256": content_hash,
            "file_size_bytes": file_info.size,
            "file_modification_ts": modified_utc,
            "filename_server_candidate": candidate_from_filename,
            "workbook_reported_server": reported_server,
            "canonical_server_name": reported_server,
            "collection_date": target_run_date,
            "collection_ts": modified_utc,
            "file_status": status,
            "validation_message": validation_message,
            "discovered_ts": now_utc,
            "processing_started_ts": None,
            "processing_completed_ts": None,
            "source_record_count": None,
            "created_ts": now_utc,
            "updated_ts": now_utc,
        }
    )

for file_info in windows_files:
    content_hash = sha256_file(file_info.path)
    current_source_file_id = source_file_id(
        "WINDOWS_EVENTS",
        file_info.path,
        file_info.modificationTime,
        file_info.size,
        content_hash,
    )
    modified_utc = modification_timestamp(file_info.modificationTime)

    file_records.append(
        {
            "source_file_id": current_source_file_id,
            "run_id": None,
            "source_type": "WINDOWS_EVENTS",
            "original_file_name": file_info.name,
            "inbox_file_path": file_info.path,
            "archived_file_path": None,
            "content_sha256": content_hash,
            "file_size_bytes": file_info.size,
            "file_modification_ts": modified_utc,
            "filename_server_candidate": None,
            "workbook_reported_server": None,
            "canonical_server_name": None,
            "collection_date": target_run_date,
            "collection_ts": modified_utc,
            "file_status": "VALIDATED",
            "validation_message": "Windows Events CSV discovered for the run date.",
            "discovered_ts": now_utc,
            "processing_started_ts": None,
            "processing_completed_ts": None,
            "source_record_count": None,
            "created_ts": now_utc,
            "updated_ts": now_utc,
        }
    )


# -----------------------------------------------------------------------------
# 4. Create a deterministic run ID from the current inbox state
# -----------------------------------------------------------------------------

run_id = f"RUN-{target_run_date:%Y%m%d}"

for record in file_records:
    record["run_id"] = run_id

print(f"Run ID: {run_id}")


# -----------------------------------------------------------------------------
# 5. Validate duplicates and the authoritative registry
# -----------------------------------------------------------------------------

valid_sql_records = [
    record
    for record in file_records
    if record["source_type"] == "SQL_DIAGNOSTICS"
    and record["file_status"] == "VALIDATED"
    and record["canonical_server_name"]
]

server_counts = Counter(
    record["canonical_server_name"] for record in valid_sql_records
)
duplicate_servers = sorted(
    server for server, count in server_counts.items() if count > 1
)

for record in valid_sql_records:
    if record["canonical_server_name"] in duplicate_servers:
        record["file_status"] = "DUPLICATE_SERVER_FILE"
        record["validation_message"] = (
            "More than one current workbook reports this canonical server."
        )

identified_servers = sorted(server_counts)

registry_rows = (
    spark.table(f"{CATALOG}.{SCHEMA}.agent_server_registry")
    .where("is_active = true AND expected_daily_workbook = true")
    .select("canonical_server_name")
    .collect()
)
registered_servers = sorted(
    {
        normalize_server_name(row["canonical_server_name"])
        for row in registry_rows
    }
)

registry_was_bootstrapped = False

if not registered_servers and BOOTSTRAP_REGISTRY:
    bootstrap_is_valid = (
        len(sql_files) == EXPECTED_SERVER_COUNT
        and len(identified_servers) == EXPECTED_SERVER_COUNT
        and not duplicate_servers
        and len(windows_files) == 1
        and all(
            record["file_status"] == "VALIDATED"
            for record in file_records
            if record["source_type"] == "SQL_DIAGNOSTICS"
        )
    )

    if not bootstrap_is_valid:
        print(
            "Registry bootstrap was requested, but the inbox does not contain "
            f"exactly {EXPECTED_SERVER_COUNT} valid and unique SQL workbooks "
            "plus one Windows Events CSV. No registry rows will be inserted."
        )
    else:
        bootstrap_rows = []

        for record in valid_sql_records:
            server_name = record["canonical_server_name"]
            bootstrap_rows.append(
                (
                    server_name,
                    server_name,
                    [record["workbook_reported_server"]],
                    [],
                    True,
                    True,
                    now_utc,
                    now_utc,
                    now_utc,
                    now_utc,
                    None,
                    "Bootstrapped from a complete validated daily workbook set.",
                )
            )

        registry_schema = StructType(
            [
                StructField("server_id", StringType(), False),
                StructField("canonical_server_name", StringType(), False),
                StructField("workbook_server_aliases", ArrayType(StringType()), False),
                StructField("windows_event_server_aliases", ArrayType(StringType()), False),
                StructField("expected_daily_workbook", BooleanType(), False),
                StructField("is_active", BooleanType(), False),
                StructField("first_seen_ts", TimestampType(), True),
                StructField("last_seen_ts", TimestampType(), True),
                StructField("created_ts", TimestampType(), True),
                StructField("updated_ts", TimestampType(), True),
                StructField("approved_by", StringType(), True),
                StructField("notes", StringType(), True),
            ]
        )

        bootstrap_df = spark.createDataFrame(bootstrap_rows, registry_schema)
        bootstrap_df.createOrReplaceTempView("_agent_registry_bootstrap")

        spark.sql(
            f"""
            MERGE INTO {table_name('agent_server_registry')} AS target
            USING _agent_registry_bootstrap AS source
               ON target.server_id = source.server_id
            WHEN NOT MATCHED THEN INSERT *
            """
        )

        spark.catalog.dropTempView("_agent_registry_bootstrap")
        registered_servers = identified_servers
        registry_was_bootstrapped = True

expected_server_set = set(registered_servers)
identified_server_set = set(identified_servers)

if registered_servers:
    missing_servers = sorted(expected_server_set - identified_server_set)
    unexpected_servers = sorted(identified_server_set - expected_server_set)
else:
    missing_servers = []
    unexpected_servers = []

for record in valid_sql_records:
    if registered_servers and record["canonical_server_name"] in unexpected_servers:
        record["file_status"] = "UNEXPECTED_SERVER"
        record["validation_message"] = (
            "Workbook server is not in the active Agent server registry."
        )


# -----------------------------------------------------------------------------
# 6. Archive valid SQL workbooks by authoritative server name
# -----------------------------------------------------------------------------

for record in file_records:
    if record["source_type"] != "SQL_DIAGNOSTICS":
        continue

    if record["file_status"] != "VALIDATED":
        continue

    server_folder = safe_path_component(record["canonical_server_name"])
    archive_directory = (
        f"{SQL_BY_SERVER}/{server_folder}/"
        f"run_date={target_run_date.isoformat()}"
    )
    archive_file_name = (
        f"{record['source_file_id'][:12]}__{record['original_file_name']}"
    )
    archive_path = f"{archive_directory}/{archive_file_name}"

    dbutils.fs.mkdirs(archive_directory)

    if not path_exists(archive_path):
        copy_succeeded = dbutils.fs.cp(
            record["inbox_file_path"],
            archive_path,
            False,
        )
        if not copy_succeeded:
            record["file_status"] = "ARCHIVE_COPY_FAILED"
            record["validation_message"] = (
                f"Could not copy workbook to {archive_path}"
            )
            continue

    record["archived_file_path"] = archive_path
    record["updated_ts"] = datetime.now(timezone.utc)


# -----------------------------------------------------------------------------
# 7. Determine final run status
# -----------------------------------------------------------------------------

invalid_workbook_count = sum(
    1
    for record in file_records
    if record["source_type"] == "SQL_DIAGNOSTICS"
    and record["file_status"] != "VALIDATED"
)

if not registered_servers:
    run_status = "REGISTRY_NOT_INITIALIZED"
elif len(windows_files) != 1:
    run_status = "INCOMPLETE_WINDOWS_INPUT"
elif duplicate_servers:
    run_status = "DUPLICATE_SERVER_INPUT"
elif missing_servers or unexpected_servers:
    run_status = "INCOMPLETE_SERVER_INPUT"
elif invalid_workbook_count:
    run_status = "INVALID_WORKBOOK_INPUT"
elif len(identified_servers) != len(registered_servers):
    run_status = "INCOMPLETE_SERVER_INPUT"
else:
    run_status = "READY_FOR_INGESTION"

if len(windows_files) > 1:
    for record in file_records:
        if record["source_type"] == "WINDOWS_EVENTS":
            record["file_status"] = "MULTIPLE_WINDOWS_FILES"
            record["validation_message"] = (
                "More than one Windows Events CSV was discovered for the run date."
            )


# -----------------------------------------------------------------------------
# 8. Idempotently persist source-file observations
# -----------------------------------------------------------------------------

source_schema = StructType(
    [
        StructField("source_file_id", StringType(), False),
        StructField("run_id", StringType(), False),
        StructField("source_type", StringType(), False),
        StructField("original_file_name", StringType(), True),
        StructField("inbox_file_path", StringType(), True),
        StructField("archived_file_path", StringType(), True),
        StructField("content_sha256", StringType(), True),
        StructField("file_size_bytes", LongType(), True),
        StructField("file_modification_ts", TimestampType(), True),
        StructField("filename_server_candidate", StringType(), True),
        StructField("workbook_reported_server", StringType(), True),
        StructField("canonical_server_name", StringType(), True),
        StructField("collection_date", DateType(), True),
        StructField("collection_ts", TimestampType(), True),
        StructField("file_status", StringType(), True),
        StructField("validation_message", StringType(), True),
        StructField("discovered_ts", TimestampType(), True),
        StructField("processing_started_ts", TimestampType(), True),
        StructField("processing_completed_ts", TimestampType(), True),
        StructField("source_record_count", LongType(), True),
        StructField("created_ts", TimestampType(), True),
        StructField("updated_ts", TimestampType(), True),
    ]
)

if file_records:
    source_df = spark.createDataFrame(file_records, source_schema)
    source_df.createOrReplaceTempView("_agent_source_file_observations")

    spark.sql(
        f"""
        MERGE INTO {table_name('agent_source_files')} AS target
        USING _agent_source_file_observations AS source
           ON target.source_file_id = source.source_file_id

        WHEN MATCHED THEN UPDATE SET
            target.run_id = source.run_id,
            target.archived_file_path = source.archived_file_path,
            target.canonical_server_name = source.canonical_server_name,
            target.file_status = source.file_status,
            target.validation_message = source.validation_message,
            target.updated_ts = source.updated_ts

        WHEN NOT MATCHED THEN INSERT *
        """
    )

    spark.catalog.dropTempView("_agent_source_file_observations")


# -----------------------------------------------------------------------------
# 9. Persist the internal run manifest and run record
# -----------------------------------------------------------------------------

manifest = {
    "run_id": run_id,
    "run_date": target_run_date.isoformat(),
    "source_timezone": SOURCE_TIMEZONE,
    "run_trigger": RUN_TRIGGER,
    "run_status": run_status,
    "expected_server_count": (
        len(registered_servers) if registered_servers else EXPECTED_SERVER_COUNT
    ),
    "discovered_workbook_count": len(sql_files),
    "identified_servers": identified_servers,
    "missing_servers": missing_servers,
    "duplicate_servers": duplicate_servers,
    "unexpected_servers": unexpected_servers,
    "windows_event_file_count": len(windows_files),
    "registry_was_bootstrapped": registry_was_bootstrapped,
    "created_ts_utc": now_utc.isoformat(),
    "source_files": [
        {
            "source_file_id": record["source_file_id"],
            "source_type": record["source_type"],
            "original_file_name": record["original_file_name"],
            "canonical_server_name": record["canonical_server_name"],
            "content_sha256": record["content_sha256"],
            "file_status": record["file_status"],
            "validation_message": record["validation_message"],
            "archived_file_path": record["archived_file_path"],
        }
        for record in file_records
    ],
}

manifest_directory = (
    f"{AGENT_ROOT}/run_manifests/run_date={target_run_date.isoformat()}"
)
manifest_path = f"{manifest_directory}/{run_id}.json"
dbutils.fs.mkdirs(manifest_directory)
dbutils.fs.put(
    manifest_path,
    json.dumps(manifest, indent=2, sort_keys=True),
    True,
)

run_error_message = None
if run_status != "READY_FOR_INGESTION":
    run_error_message = (
        f"Daily input validation did not reach READY_FOR_INGESTION. "
        f"Status={run_status}. Manifest={manifest_path}"
    )

run_rows = [
    (
        run_id,
        target_run_date,
        SOURCE_TIMEZONE,
        run_status,
        RUN_TRIGGER,
        len(registered_servers) if registered_servers else EXPECTED_SERVER_COUNT,
        len(sql_files),
        len(identified_servers),
        sum(
            1
            for record in file_records
            if record["source_type"] == "SQL_DIAGNOSTICS"
            and record["file_status"] == "VALIDATED"
        ),
        invalid_workbook_count,
        len(windows_files),
        missing_servers,
        duplicate_servers,
        unexpected_servers,
        now_utc,
        datetime.now(timezone.utc),
        None,
        run_error_message,
        now_utc,
        datetime.now(timezone.utc),
    )
]

run_schema = StructType(
    [
        StructField("run_id", StringType(), False),
        StructField("run_date", DateType(), False),
        StructField("source_timezone", StringType(), False),
        StructField("run_status", StringType(), False),
        StructField("run_trigger", StringType(), False),
        StructField("expected_server_count", IntegerType(), False),
        StructField("discovered_workbook_count", IntegerType(), False),
        StructField("identified_server_count", IntegerType(), False),
        StructField("valid_workbook_count", IntegerType(), False),
        StructField("invalid_workbook_count", IntegerType(), False),
        StructField("windows_event_file_count", IntegerType(), False),
        StructField("missing_servers", ArrayType(StringType()), False),
        StructField("duplicate_servers", ArrayType(StringType()), False),
        StructField("unexpected_servers", ArrayType(StringType()), False),
        StructField("started_ts", TimestampType(), True),
        StructField("validation_completed_ts", TimestampType(), True),
        StructField("processing_completed_ts", TimestampType(), True),
        StructField("error_message", StringType(), True),
        StructField("created_ts", TimestampType(), True),
        StructField("updated_ts", TimestampType(), True),
    ]
)

run_df = spark.createDataFrame(run_rows, run_schema)
run_df.createOrReplaceTempView("_agent_ingestion_run")

spark.sql(
    f"""
    MERGE INTO {table_name('agent_ingestion_runs')} AS target
    USING _agent_ingestion_run AS source
       ON target.run_id = source.run_id

    WHEN MATCHED THEN UPDATE SET
        target.run_status = source.run_status,
        target.discovered_workbook_count = source.discovered_workbook_count,
        target.identified_server_count = source.identified_server_count,
        target.valid_workbook_count = source.valid_workbook_count,
        target.invalid_workbook_count = source.invalid_workbook_count,
        target.windows_event_file_count = source.windows_event_file_count,
        target.missing_servers = source.missing_servers,
        target.duplicate_servers = source.duplicate_servers,
        target.unexpected_servers = source.unexpected_servers,
        target.validation_completed_ts = source.validation_completed_ts,
        target.error_message = source.error_message,
        target.updated_ts = source.updated_ts

    WHEN NOT MATCHED THEN INSERT *
    """
)

spark.catalog.dropTempView("_agent_ingestion_run")


# -----------------------------------------------------------------------------
# 10. Output for manual validation and downstream Job tasks
# -----------------------------------------------------------------------------

summary_rows = [
    ("run_id", run_id),
    ("run_date", target_run_date.isoformat()),
    ("run_status", run_status),
    ("expected_server_count", str(len(registered_servers) or EXPECTED_SERVER_COUNT)),
    ("discovered_workbook_count", str(len(sql_files))),
    ("identified_server_count", str(len(identified_servers))),
    ("invalid_workbook_count", str(invalid_workbook_count)),
    ("windows_event_file_count", str(len(windows_files))),
    ("missing_server_count", str(len(missing_servers))),
    ("duplicate_server_count", str(len(duplicate_servers))),
    ("unexpected_server_count", str(len(unexpected_servers))),
    ("manifest_path", manifest_path),
]

display(spark.createDataFrame(summary_rows, ["check", "value"]))

if file_records:
    display(
        spark.createDataFrame(file_records, source_schema).select(
            "source_type",
            "original_file_name",
            "canonical_server_name",
            "file_status",
            "validation_message",
            "archived_file_path",
        )
    )

try:
    dbutils.jobs.taskValues.set(key="run_id", value=run_id)
    dbutils.jobs.taskValues.set(key="run_status", value=run_status)
    dbutils.jobs.taskValues.set(key="manifest_path", value=manifest_path)
except Exception:
    # Expected when the notebook is run interactively rather than as a Job task.
    pass

print(f"Validation completed with status: {run_status}")
print(f"Internal manifest: {manifest_path}")

if FAIL_ON_INCOMPLETE and run_status != "READY_FOR_INGESTION":
    raise RuntimeError(run_error_message)