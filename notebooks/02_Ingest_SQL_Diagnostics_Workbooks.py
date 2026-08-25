# Databricks notebook source
# 02_Ingest_SQL_Diagnostics_Workbooks

from __future__ import annotations

import json
import math
import os
import re
from datetime import date, datetime, time, timezone
from decimal import Decimal
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from pyspark.sql import functions as F
from pyspark.sql.window import Window
from pyspark.sql.types import (
    BooleanType,
    DateType,
    IntegerType,
    LongType,
    StringType,
    StructField,
    StructType,
    TimestampType,
)


# -----------------------------------------------------------------------------
# 1. Runtime parameters
# -----------------------------------------------------------------------------

dbutils.widgets.text("run_id", "", "Run ID (blank = current business date)")
dbutils.widgets.dropdown(
    "allow_incomplete_run",
    "false",
    ["false", "true"],
    "Allow incomplete run for development testing",
)
dbutils.widgets.dropdown(
    "force_reprocess",
    "false",
    ["false", "true"],
    "Force reprocessing of already ingested files",
)
dbutils.widgets.dropdown(
    "fail_on_workbook_error",
    "true",
    ["true", "false"],
    "Fail task when any workbook cannot be fully ingested",
)

CATALOG = "ent_log_analytics"
SCHEMA = "observability"


def table_name(name: str) -> str:
    return f"`{CATALOG}`.`{SCHEMA}`.`{name}`"


CONFIG = {
    row["config_key"]: row["config_value"]
    for row in spark.table(f"{CATALOG}.{SCHEMA}.agent_config")
    .select("config_key", "config_value")
    .collect()
}

SOURCE_TIMEZONE = CONFIG.get("source_timezone", "America/New_York")
AGENT_ROOT = CONFIG["agent_root_path"]

run_id_parameter = dbutils.widgets.get("run_id").strip()
ALLOW_INCOMPLETE_RUN = (
    dbutils.widgets.get("allow_incomplete_run").strip().lower() == "true"
)
FORCE_REPROCESS = (
    dbutils.widgets.get("force_reprocess").strip().lower() == "true"
)
FAIL_ON_WORKBOOK_ERROR = (
    dbutils.widgets.get("fail_on_workbook_error").strip().lower() == "true"
)

if run_id_parameter:
    RUN_ID = run_id_parameter
else:
    latest_run_rows = (
        spark.table(f"{CATALOG}.{SCHEMA}.agent_ingestion_runs")
        .orderBy(F.col("run_date").desc(), F.col("updated_ts").desc())
        .select("run_id")
        .limit(1)
        .collect()
    )
    if not latest_run_rows:
        raise RuntimeError(
            "No Agent ingestion run exists. Run notebook 01 first."
        )
    RUN_ID = latest_run_rows[0]["run_id"]

run_rows = (
    spark.table(f"{CATALOG}.{SCHEMA}.agent_ingestion_runs")
    .where(F.col("run_id") == RUN_ID)
    .orderBy(F.col("updated_ts").desc())
    .limit(1)
    .collect()
)

if not run_rows:
    raise RuntimeError(f"Run ID was not found: {RUN_ID}")

run_row = run_rows[0]
RUN_DATE = run_row["run_date"]
VALIDATION_STATUS = run_row["run_status"]

production_ready_statuses = {
    "READY_FOR_INGESTION",
    "SQL_DIAGNOSTICS_INGESTED",
}

if VALIDATION_STATUS not in production_ready_statuses and not ALLOW_INCOMPLETE_RUN:
    raise RuntimeError(
        f"Run {RUN_ID} has status {VALIDATION_STATUS}. "
        "Production ingestion requires READY_FOR_INGESTION. "
        "Use allow_incomplete_run=true only for a controlled development test."
    )

print(f"Run ID: {RUN_ID}")
print(f"Run date: {RUN_DATE}")
print(f"Validation status: {VALIDATION_STATUS}")
print(f"Allow incomplete run: {ALLOW_INCOMPLETE_RUN}")
print(f"Force reprocess: {FORCE_REPROCESS}")


# -----------------------------------------------------------------------------
# 2. Verified diagnostic workbook contract
# -----------------------------------------------------------------------------

# This list is taken from the representative production workbook.
# It is used to detect missing collection sections, not to limit ingestion.
# Additional worksheets are still ingested and identified as unexpected.
EXPECTED_SHEETS = [
    "1-Version Info",
    "2-Core Counts",
    "3-Server Properties",
    "4-Configuration Values",
    "5-Global Trace Flags",
    "6-Process Memory",
    "7-SQL Server Services Info",
    "8-Last Backup By Database",
    "9-Accelerator Status",
    "10-SQL Server Agent Jobs",
    "11-SQL Server Agent Alerts",
    "12-Host Info",
    "13-SQL Server NUMA Info",
    "14-System Memory",
    "15-AlwaysOn AG Cluster",
    "16-AG Status",
    "17-Hardware Info",
    "18-System Manufacturer",
    "19-BIOS Date",
    "20-Processor Description",
    "21-CPU Vectorization Level",
    "22-Memory Dump Info",
    "23-Suspect Pages",
    "24-TempDB Data Files",
    "25-Tempdb Data File Sizes",
    "26-Database Filenames and Path",
    "27-Fixed Drives",
    "28-Volume Info",
    "29-Drive Level Latency",
    "30-IO Latency by File",
    "31-IO Warnings",
    "32-RG Resource Pools",
    "33-Database Properties",
    "34-Missing Indexes All Databas",
    "35-VLF Counts",
    "36-CPU Usage by Database",
    "37-IO Usage By Database",
    "38-Total Buffer Usage by Datab",
    "39-Version Store Space Usage",
    "40-Top Waits",
    "41-Connection Counts by IP Add",
    "42-Avg Task Counts",
    "43-Detect Blocking",
    "44-Page Contention",
    "45-CPU Utilization History",
    "46-Top Worker Time Queries",
    "47-PLE by NUMA Node",
    "48-Memory Grants Pending",
    "49-Memory Clerk Usage",
    "50-Ad hoc Queries",
    "51-Top Logical Reads Queries",
    "52-Top Avg Elapsed Time Querie",
    "53-UDF Stats by DB",
]

EXPECTED_SHEET_SET = set(EXPECTED_SHEETS)
EXPECTED_ORDINAL_BY_NAME = {
    name: ordinal for ordinal, name in enumerate(EXPECTED_SHEETS, start=1)
}

MAX_ROWS_PER_SHEET = 2_000_000
MAX_COLUMNS_PER_SHEET = 500


# -----------------------------------------------------------------------------
# 3. Schemas and serialization helpers
# -----------------------------------------------------------------------------

BRONZE_SCHEMA = StructType(
    [
        StructField("run_id", StringType(), False),
        StructField("source_file_id", StringType(), False),
        StructField("canonical_server_name", StringType(), False),
        StructField("workbook_reported_server", StringType(), True),
        StructField("snapshot_date", DateType(), False),
        StructField("ingestion_date", DateType(), False),
        StructField("sheet_ordinal", IntegerType(), False),
        StructField("sheet_name", StringType(), False),
        StructField("source_row_number", LongType(), False),
        StructField("row_json", StringType(), False),
        StructField("source_file_path", StringType(), False),
        StructField("ingested_ts", TimestampType(), False),
    ]
)

MANIFEST_SCHEMA = StructType(
    [
        StructField("run_id", StringType(), False),
        StructField("source_file_id", StringType(), False),
        StructField("canonical_server_name", StringType(), False),
        StructField("sheet_ordinal", IntegerType(), False),
        StructField("sheet_name", StringType(), False),
        StructField("sheet_status", StringType(), False),
        StructField("no_data_marker_found", BooleanType(), False),
        StructField("source_row_count", LongType(), False),
        StructField("ingested_row_count", LongType(), False),
        StructField("detected_columns_json", StringType(), True),
        StructField("processing_started_ts", TimestampType(), True),
        StructField("processing_completed_ts", TimestampType(), True),
        StructField("error_message", StringType(), True),
        StructField("created_ts", TimestampType(), False),
    ]
)


def to_fuse_path(path: str) -> str:
    if path.startswith("dbfs:/Volumes/"):
        return path[len("dbfs:"):]
    return path


def normalize_server_name(value: Any) -> str:
    value_text = "" if value is None else str(value)
    return re.sub(r"\s+", "", value_text).strip().upper()


def normalized_marker(value: Any) -> str:
    return re.sub(
        r"[^a-z0-9]",
        "",
        "" if value is None else str(value).strip().lower(),
    )


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def row_is_blank(values: Iterable[Any]) -> bool:
    return all(is_blank(value) for value in values)


def json_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, bytes):
        return value.hex()
    if isinstance(value, float):
        if math.isnan(value) or math.isinf(value):
            return None
        return value
    if isinstance(value, (str, int, bool)):
        return value
    return str(value)


def unique_headers(header_values: Iterable[Any]) -> list[str]:
    result = []
    counts: dict[str, int] = {}

    for column_number, value in enumerate(header_values, start=1):
        base_name = str(value).strip() if not is_blank(value) else f"_column_{column_number}"
        occurrence = counts.get(base_name, 0) + 1
        counts[base_name] = occurrence
        unique_name = base_name if occurrence == 1 else f"{base_name}__{occurrence}"
        result.append(unique_name)

    return result


def copy_to_quarantine(source_path: str, source_file_id: str) -> str | None:
    quarantine_directory = (
        f"{AGENT_ROOT}/quarantine/sql_diagnostics/"
        f"run_date={RUN_DATE.isoformat()}"
    )
    quarantine_path = (
        f"{quarantine_directory}/{source_file_id[:12]}__"
        f"{Path(source_path).name}"
    )

    try:
        dbutils.fs.mkdirs(quarantine_directory)
        if not os.path.exists(to_fuse_path(quarantine_path)):
            copied = dbutils.fs.cp(source_path, quarantine_path, False)
            if not copied:
                return None
        return quarantine_path
    except Exception:
        return None


def set_source_status(
    source_file_id: str,
    status: str,
    message: str | None,
    row_count: int | None,
    started_ts: datetime | None,
    completed_ts: datetime | None,
) -> None:
    update_rows = [
        (
            source_file_id,
            status,
            message,
            row_count,
            started_ts,
            completed_ts,
            datetime.now(timezone.utc),
        )
    ]
    update_schema = StructType(
        [
            StructField("source_file_id", StringType(), False),
            StructField("file_status", StringType(), False),
            StructField("validation_message", StringType(), True),
            StructField("source_record_count", LongType(), True),
            StructField("processing_started_ts", TimestampType(), True),
            StructField("processing_completed_ts", TimestampType(), True),
            StructField("updated_ts", TimestampType(), False),
        ]
    )
    spark.createDataFrame(update_rows, update_schema).createOrReplaceTempView(
        "_agent_source_status_update"
    )
    spark.sql(
        f"""
        MERGE INTO {table_name('agent_source_files')} AS target
        USING _agent_source_status_update AS source
           ON target.source_file_id = source.source_file_id
        WHEN MATCHED THEN UPDATE SET
            target.file_status = source.file_status,
            target.validation_message = source.validation_message,
            target.source_record_count = source.source_record_count,
            target.processing_started_ts = source.processing_started_ts,
            target.processing_completed_ts = source.processing_completed_ts,
            target.updated_ts = source.updated_ts
        """
    )
    spark.catalog.dropTempView("_agent_source_status_update")


def merge_manifest(source_file_id: str, rows: list[tuple[Any, ...]]) -> None:
    manifest_df = spark.createDataFrame(rows, MANIFEST_SCHEMA)
    manifest_df.createOrReplaceTempView("_agent_workbook_sheet_manifest")

    spark.sql(
        f"""
        MERGE INTO {table_name('agent_sheet_manifest')} AS target
        USING _agent_workbook_sheet_manifest AS source
           ON target.source_file_id = source.source_file_id
          AND target.sheet_ordinal = source.sheet_ordinal
          AND target.sheet_name = source.sheet_name

        WHEN MATCHED THEN UPDATE SET
            target.run_id = source.run_id,
            target.canonical_server_name = source.canonical_server_name,
            target.sheet_status = source.sheet_status,
            target.no_data_marker_found = source.no_data_marker_found,
            target.source_row_count = source.source_row_count,
            target.ingested_row_count = source.ingested_row_count,
            target.detected_columns_json = source.detected_columns_json,
            target.processing_started_ts = source.processing_started_ts,
            target.processing_completed_ts = source.processing_completed_ts,
            target.error_message = source.error_message

        WHEN NOT MATCHED THEN INSERT *

        WHEN NOT MATCHED BY SOURCE
          AND target.source_file_id = '{source_file_id}'
        THEN DELETE
        """
    )

    spark.catalog.dropTempView("_agent_workbook_sheet_manifest")


def merge_bronze(
    source_file_id: str,
    rows: list[tuple[Any, ...]],
) -> None:
    if not rows:
        spark.sql(
            f"""
            DELETE FROM {table_name('agent_sql_diagnostics_bronze')}
            WHERE source_file_id = '{source_file_id}'
            """
        )
        return

    bronze_df = spark.createDataFrame(rows, BRONZE_SCHEMA)
    bronze_df.createOrReplaceTempView("_agent_workbook_bronze_rows")

    spark.sql(
        f"""
        MERGE INTO {table_name('agent_sql_diagnostics_bronze')} AS target
        USING _agent_workbook_bronze_rows AS source
           ON target.source_file_id = source.source_file_id
          AND target.sheet_ordinal = source.sheet_ordinal
          AND target.source_row_number = source.source_row_number

        WHEN MATCHED THEN UPDATE SET
            target.run_id = source.run_id,
            target.canonical_server_name = source.canonical_server_name,
            target.workbook_reported_server = source.workbook_reported_server,
            target.snapshot_date = source.snapshot_date,
            target.ingestion_date = source.ingestion_date,
            target.sheet_name = source.sheet_name,
            target.row_json = source.row_json,
            target.source_file_path = source.source_file_path,
            target.ingested_ts = source.ingested_ts

        WHEN NOT MATCHED THEN INSERT *

        WHEN NOT MATCHED BY SOURCE
          AND target.source_file_id = '{source_file_id}'
        THEN DELETE
        """
    )

    spark.catalog.dropTempView("_agent_workbook_bronze_rows")


# -----------------------------------------------------------------------------
# 4. Select the latest validated workbook for each server in the run
# -----------------------------------------------------------------------------

eligible_statuses = ["VALIDATED", "INGESTED", "INGESTION_FAILED"]

candidate_files = (
    spark.table(f"{CATALOG}.{SCHEMA}.agent_source_files")
    .where(
        (F.col("run_id") == RUN_ID)
        & (F.col("source_type") == "SQL_DIAGNOSTICS")
        & F.col("canonical_server_name").isNotNull()
        & F.col("file_status").isin(eligible_statuses)
    )
)

latest_file_window = Window.partitionBy("canonical_server_name").orderBy(
    F.col("file_modification_ts").desc(),
    F.col("discovered_ts").desc(),
    F.col("source_file_id").desc(),
)

selected_files = (
    candidate_files
    .withColumn("_latest_rank", F.row_number().over(latest_file_window))
    .where(F.col("_latest_rank") == 1)
    .drop("_latest_rank")
    .orderBy("canonical_server_name")
    .collect()
)

if not selected_files:
    raise RuntimeError(
        f"No validated SQL diagnostic workbooks were found for {RUN_ID}."
    )

print(f"Latest server workbooks selected: {len(selected_files)}")


# -----------------------------------------------------------------------------
# 5. Parse, validate and ingest each workbook
# -----------------------------------------------------------------------------

file_results: list[tuple[str, str, int, int, str | None]] = []
failed_files: list[tuple[str, str]] = []

for source in selected_files:
    source_file_id = source["source_file_id"]
    server_name = normalize_server_name(source["canonical_server_name"])
    source_path = source["archived_file_path"] or source["inbox_file_path"]
    workbook_reported_server = normalize_server_name(
        source["workbook_reported_server"]
    )

    if source["file_status"] == "INGESTED" and not FORCE_REPROCESS:
        existing_count = int(source["source_record_count"] or 0)
        file_results.append(
            (server_name, "SKIPPED_ALREADY_INGESTED", 0, existing_count, None)
        )
        print(f"Skipped already ingested workbook: {server_name}")
        continue

    processing_started = datetime.now(timezone.utc)
    set_source_status(
        source_file_id,
        "PROCESSING",
        "SQL diagnostic workbook ingestion started.",
        None,
        processing_started,
        None,
    )

    workbook = None
    bronze_rows: list[tuple[Any, ...]] = []
    manifest_rows: list[tuple[Any, ...]] = []
    workbook_errors: list[str] = []

    try:
        if not source_path or not os.path.exists(to_fuse_path(source_path)):
            raise FileNotFoundError(
                f"Validated workbook is not accessible: {source_path}"
            )

        workbook = load_workbook(
            filename=to_fuse_path(source_path),
            read_only=True,
            data_only=True,
        )

        actual_sheet_names = list(workbook.sheetnames)
        if not actual_sheet_names:
            raise ValueError("Workbook contains no worksheets")

        first_sheet = workbook[actual_sheet_names[0]]
        actual_reported_server = normalize_server_name(first_sheet["A2"].value)

        if actual_reported_server != server_name:
            raise ValueError(
                "Workbook identity changed after validation: "
                f"registered={server_name}, workbook={actual_reported_server}"
            )

        missing_expected_sheets = [
            sheet_name
            for sheet_name in EXPECTED_SHEETS
            if sheet_name not in actual_sheet_names
        ]

        for missing_sheet in missing_expected_sheets:
            workbook_errors.append(
                f"Missing expected worksheet: {missing_sheet}"
            )
            ordinal = EXPECTED_ORDINAL_BY_NAME[missing_sheet]
            completed_ts = datetime.now(timezone.utc)
            manifest_rows.append(
                (
                    RUN_ID,
                    source_file_id,
                    server_name,
                    ordinal,
                    missing_sheet,
                    "MISSING_EXPECTED",
                    False,
                    0,
                    0,
                    None,
                    processing_started,
                    completed_ts,
                    "Expected worksheet is absent from the workbook.",
                    processing_started,
                )
            )

        for sheet_ordinal, sheet_name in enumerate(actual_sheet_names, start=1):
            sheet_started = datetime.now(timezone.utc)
            worksheet = workbook[sheet_name]
            is_expected_sheet = sheet_name in EXPECTED_SHEET_SET

            try:
                if worksheet.max_row > MAX_ROWS_PER_SHEET:
                    raise ValueError(
                        f"Worksheet has {worksheet.max_row} rows; "
                        f"limit is {MAX_ROWS_PER_SHEET}."
                    )
                if worksheet.max_column > MAX_COLUMNS_PER_SHEET:
                    raise ValueError(
                        f"Worksheet has {worksheet.max_column} columns; "
                        f"limit is {MAX_COLUMNS_PER_SHEET}."
                    )

                row_iterator = worksheet.iter_rows(values_only=True)
                header_row = next(row_iterator, None)

                if header_row is None or row_is_blank(header_row):
                    status = "EMPTY" if is_expected_sheet else "EMPTY_UNEXPECTED"
                    if is_expected_sheet:
                        workbook_errors.append(
                            f"Expected worksheet is empty: {sheet_name}"
                        )
                    manifest_rows.append(
                        (
                            RUN_ID,
                            source_file_id,
                            server_name,
                            sheet_ordinal,
                            sheet_name,
                            status,
                            False,
                            0,
                            0,
                            None,
                            sheet_started,
                            datetime.now(timezone.utc),
                            None,
                            processing_started,
                        )
                    )
                    continue

                first_nonblank_value = next(
                    (value for value in header_row if not is_blank(value)),
                    None,
                )

                if normalized_marker(first_nonblank_value) == "nodata":
                    status = "NO_DATA" if is_expected_sheet else "NO_DATA_UNEXPECTED"
                    manifest_rows.append(
                        (
                            RUN_ID,
                            source_file_id,
                            server_name,
                            sheet_ordinal,
                            sheet_name,
                            status,
                            True,
                            0,
                            0,
                            json.dumps([str(first_nonblank_value)]),
                            sheet_started,
                            datetime.now(timezone.utc),
                            None,
                            processing_started,
                        )
                    )
                    continue

                headers = unique_headers(header_row)
                source_row_count = 0
                sheet_bronze_rows: list[tuple[Any, ...]] = []
                ingested_ts = datetime.now(timezone.utc)

                for excel_row_number, row_values in enumerate(
                    row_iterator,
                    start=2,
                ):
                    if row_is_blank(row_values):
                        continue

                    padded_values = list(row_values[: len(headers)])
                    if len(padded_values) < len(headers):
                        padded_values.extend(
                            [None] * (len(headers) - len(padded_values))
                        )

                    row_payload = {
                        header: json_value(value)
                        for header, value in zip(headers, padded_values)
                    }

                    row_json = json.dumps(
                        row_payload,
                        ensure_ascii=False,
                        separators=(",", ":"),
                    )

                    sheet_bronze_rows.append(
                        (
                            RUN_ID,
                            source_file_id,
                            server_name,
                            workbook_reported_server,
                            RUN_DATE,
                            RUN_DATE,
                            sheet_ordinal,
                            sheet_name,
                            excel_row_number,
                            row_json,
                            source_path,
                            ingested_ts,
                        )
                    )
                    source_row_count += 1

                if source_row_count == 0:
                    status = "EMPTY" if is_expected_sheet else "EMPTY_UNEXPECTED"
                    if is_expected_sheet:
                        workbook_errors.append(
                            f"Expected worksheet has headers but no rows: {sheet_name}"
                        )
                else:
                    status = "INGESTED" if is_expected_sheet else "INGESTED_UNEXPECTED"
                    bronze_rows.extend(sheet_bronze_rows)

                manifest_rows.append(
                    (
                        RUN_ID,
                        source_file_id,
                        server_name,
                        sheet_ordinal,
                        sheet_name,
                        status,
                        False,
                        source_row_count,
                        source_row_count,
                        json.dumps(headers, ensure_ascii=False),
                        sheet_started,
                        datetime.now(timezone.utc),
                        None,
                        processing_started,
                    )
                )

            except Exception as sheet_exc:
                workbook_errors.append(
                    f"Worksheet {sheet_name} failed: {sheet_exc}"
                )
                manifest_rows.append(
                    (
                        RUN_ID,
                        source_file_id,
                        server_name,
                        sheet_ordinal,
                        sheet_name,
                        "READ_ERROR",
                        False,
                        0,
                        0,
                        None,
                        sheet_started,
                        datetime.now(timezone.utc),
                        str(sheet_exc),
                        processing_started,
                    )
                )

        if workbook_errors:
            # No Bronze rows from a structurally incomplete workbook are committed.
            for index, manifest_row in enumerate(manifest_rows):
                row_list = list(manifest_row)
                if row_list[5] in {"INGESTED", "INGESTED_UNEXPECTED"}:
                    row_list[5] = "PARSED_NOT_COMMITTED"
                    row_list[8] = 0
                manifest_rows[index] = tuple(row_list)

            merge_manifest(source_file_id, manifest_rows)
            quarantine_path = copy_to_quarantine(source_path, source_file_id)
            error_message = " | ".join(workbook_errors)[:10000]
            if quarantine_path:
                error_message += f" | quarantine={quarantine_path}"

            set_source_status(
                source_file_id,
                "INGESTION_FAILED",
                error_message,
                0,
                processing_started,
                datetime.now(timezone.utc),
            )
            failed_files.append((server_name, error_message))
            file_results.append(
                (server_name, "INGESTION_FAILED", len(actual_sheet_names), 0, error_message)
            )
            continue

        merge_bronze(source_file_id, bronze_rows)
        merge_manifest(source_file_id, manifest_rows)

        total_rows = len(bronze_rows)
        set_source_status(
            source_file_id,
            "INGESTED",
            (
                f"Successfully ingested all {len(actual_sheet_names)} worksheets; "
                f"Bronze rows={total_rows}."
            ),
            total_rows,
            processing_started,
            datetime.now(timezone.utc),
        )
        file_results.append(
            (server_name, "INGESTED", len(actual_sheet_names), total_rows, None)
        )
        print(
            f"Ingested {server_name}: sheets={len(actual_sheet_names)}, "
            f"rows={total_rows}"
        )

    except Exception as workbook_exc:
        error_message = str(workbook_exc)
        quarantine_path = copy_to_quarantine(source_path, source_file_id)
        if quarantine_path:
            error_message += f" | quarantine={quarantine_path}"

        set_source_status(
            source_file_id,
            "INGESTION_FAILED",
            error_message,
            0,
            processing_started,
            datetime.now(timezone.utc),
        )
        failed_files.append((server_name, error_message))
        file_results.append(
            (server_name, "INGESTION_FAILED", 0, 0, error_message)
        )

    finally:
        if workbook is not None:
            workbook.close()


# -----------------------------------------------------------------------------
# 6. Latest Agent SQL diagnostic views
# -----------------------------------------------------------------------------

spark.sql(
    f"""
    CREATE OR REPLACE VIEW {table_name('v_agent_latest_sql_source_files')} AS
    SELECT
        source_file_id,
        run_id,
        canonical_server_name,
        workbook_reported_server,
        original_file_name,
        inbox_file_path,
        archived_file_path,
        content_sha256,
        file_size_bytes,
        file_modification_ts,
        collection_date,
        collection_ts,
        file_status,
        source_record_count,
        processing_completed_ts
    FROM (
        SELECT
            source_file_id,
            run_id,
            canonical_server_name,
            workbook_reported_server,
            original_file_name,
            inbox_file_path,
            archived_file_path,
            content_sha256,
            file_size_bytes,
            file_modification_ts,
            collection_date,
            collection_ts,
            file_status,
            source_record_count,
            processing_completed_ts,
            ROW_NUMBER() OVER (
                PARTITION BY run_id, canonical_server_name
                ORDER BY file_modification_ts DESC,
                         processing_completed_ts DESC,
                         source_file_id DESC
            ) AS latest_rank
        FROM {table_name('agent_source_files')}
        WHERE source_type = 'SQL_DIAGNOSTICS'
          AND file_status = 'INGESTED'
    )
    WHERE latest_rank = 1
    """
)

spark.sql(
    f"""
    CREATE OR REPLACE VIEW {table_name('v_agent_latest_sql_diagnostics')} AS
    SELECT bronze.*
    FROM {table_name('agent_sql_diagnostics_bronze')} AS bronze
    INNER JOIN {table_name('v_agent_latest_sql_source_files')} AS latest
        ON bronze.source_file_id = latest.source_file_id
       AND bronze.run_id = latest.run_id
    """
)


# -----------------------------------------------------------------------------
# 7. Update run status and display validation results
# -----------------------------------------------------------------------------

if failed_files:
    final_run_status = "SQL_DIAGNOSTICS_INGESTION_FAILED"
    final_error_message = " | ".join(
        f"{server}: {message}" for server, message in failed_files
    )[:10000]
elif VALIDATION_STATUS in production_ready_statuses:
    final_run_status = "SQL_DIAGNOSTICS_INGESTED"
    final_error_message = None
else:
    final_run_status = "TEST_SQL_DIAGNOSTICS_INGESTED"
    final_error_message = (
        "SQL ingestion completed under allow_incomplete_run=true for development testing."
    )

run_update_schema = StructType(
    [
        StructField("run_id", StringType(), False),
        StructField("run_status", StringType(), False),
        StructField("processing_completed_ts", TimestampType(), True),
        StructField("error_message", StringType(), True),
        StructField("updated_ts", TimestampType(), False),
    ]
)

spark.createDataFrame(
    [
        (
            RUN_ID,
            final_run_status,
            datetime.now(timezone.utc),
            final_error_message,
            datetime.now(timezone.utc),
        )
    ],
    run_update_schema,
).createOrReplaceTempView("_agent_sql_ingestion_run_update")

spark.sql(
    f"""
    MERGE INTO {table_name('agent_ingestion_runs')} AS target
    USING _agent_sql_ingestion_run_update AS source
       ON target.run_id = source.run_id
    WHEN MATCHED THEN UPDATE SET
        target.run_status = source.run_status,
        target.processing_completed_ts = source.processing_completed_ts,
        target.error_message = source.error_message,
        target.updated_ts = source.updated_ts
    """
)

spark.catalog.dropTempView("_agent_sql_ingestion_run_update")

result_schema = StructType(
    [
        StructField("canonical_server_name", StringType(), False),
        StructField("result", StringType(), False),
        StructField("worksheet_count", IntegerType(), False),
        StructField("bronze_row_count", LongType(), False),
        StructField("error_message", StringType(), True),
    ]
)

display(spark.createDataFrame(file_results, result_schema))

display(
    spark.table(f"{CATALOG}.{SCHEMA}.agent_sheet_manifest")
    .where(F.col("run_id") == RUN_ID)
    .groupBy("canonical_server_name", "sheet_status")
    .agg(
        F.count("*").alias("worksheet_count"),
        F.sum("source_row_count").alias("source_rows"),
        F.sum("ingested_row_count").alias("ingested_rows"),
    )
    .orderBy("canonical_server_name", "sheet_status")
)

print(f"Final run status: {final_run_status}")

try:
    dbutils.jobs.taskValues.set(key="run_id", value=RUN_ID)
    dbutils.jobs.taskValues.set(key="sql_ingestion_status", value=final_run_status)
except Exception:
    pass

if failed_files and FAIL_ON_WORKBOOK_ERROR:
    raise RuntimeError(final_error_message)